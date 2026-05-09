#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
PEC / IMAP backup tool with a PySide6 interface.

Goals:
- search IMAP messages older than a cutoff date;
- save every message as a RAW .eml file, byte-for-byte, including text,
  attachments, MIME structure and signatures, like Thunderbird "Save As";
- create/update an Excel .xlsx index;
- avoid duplicates with SHA-256 and a short hash in the file name;
- for PEC messages, try daticert.xml / postacert.eml to find the real
  counterparty instead of generic senders such as posta-certificata@...;
- set \Deleted only at the end, after verified backup and explicit confirmation.

Dependencies:
    pip install PySide6 openpyxl

Password safety note:
- when "save password" is enabled, the password is stored locally in QSettings
  using a simple base64/XOR obfuscation. This is not strong encryption.
"""

from __future__ import annotations

import base64
import dataclasses
import email
import email.policy
import faulthandler
import hashlib
import imaplib
import json
import os
import re
import socket
import ssl
import sys

# ---------------------------------------------------------------------------
# Very early crash/Qt debug: keep this before importing PySide6.
# ---------------------------------------------------------------------------

EARLY_ARGS = set(sys.argv[1:])
SAFE_MODE = "--safe-mode" in EARLY_ARGS
RESET_SETTINGS_REQUESTED = "--reset-settings" in EARLY_ARGS
RESTORE_UI_STATE = "--restore-ui-state" in EARLY_ARGS and not SAFE_MODE
QT_DEBUG_PLUGINS_REQUESTED = "--qt-debug-plugins" in EARLY_ARGS

if QT_DEBUG_PLUGINS_REQUESTED:
    os.environ.setdefault("QT_DEBUG_PLUGINS", "1")

if "--force-xcb" in EARLY_ARGS:
    os.environ["QT_QPA_PLATFORM"] = "xcb"
elif "--force-wayland" in EARLY_ARGS:
    os.environ["QT_QPA_PLATFORM"] = "wayland"

if SAFE_MODE:
    # Avoid common segfault sources in Qt/graphics drivers and do not restore
    # binary Qt states saved by previous script versions.
    os.environ.setdefault("QT_OPENGL", "software")
    os.environ.setdefault("LIBGL_ALWAYS_SOFTWARE", "1")

_CRASH_LOG_FILE = None
try:
    crash_log_path = os.environ.get("PEC_IMAP_BACKUP_CRASH_LOG") or os.path.join(os.getcwd(), "pec_imap_backup_crash.log")
    _CRASH_LOG_FILE = open(crash_log_path, "a", encoding="utf-8", buffering=1)
    _CRASH_LOG_FILE.write("\n--- avvio imap_backup.py ---\n")
    _CRASH_LOG_FILE.write("argv=" + repr(sys.argv) + "\n")
    _CRASH_LOG_FILE.write("QT_QPA_PLATFORM=" + repr(os.environ.get("QT_QPA_PLATFORM")) + "\n")
    faulthandler.enable(file=_CRASH_LOG_FILE, all_threads=True)
except Exception:
    try:
        faulthandler.enable(all_threads=True)
    except Exception:
        pass

def clean_qt_argv(argv):
    custom = {
        "--safe-mode",
        "--reset-settings",
        "--restore-ui-state",
        "--qt-debug-plugins",
        "--force-xcb",
        "--force-wayland",
        "--debug-help",
    }
    return [argv[0], *[a for a in argv[1:] if a not in custom]]

import tempfile
import threading
import traceback
import unicodedata
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from email.header import decode_header, make_header
from email.message import Message
from email.utils import getaddresses, parseaddr, parsedate_to_datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from PySide6.QtCore import QObject, QSettings, QThread, Qt, QDate, QTimer, Signal, Slot, qInstallMessageHandler
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QProgressBar,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

APP_ORG = "LocalTools"
APP_NAME = "PEC IMAP Backup"
INDEX_FILENAME = "indice_backup_pec.xlsx"

INDEX_HEADERS = [
    "saved_at",
    "status",
    "deleted_at",
    "folder",
    "uid",
    "message_id",
    "date_header",
    "internal_date",
    "direction",
    "mittente_vero",
    "header_from",
    "header_to",
    "subject_header",
    "pec_type",
    "pec_party",
    "pec_mittente",
    "pec_destinatari",
    "pec_oggetto",
    "filename",
    "relative_path",
    "size_bytes",
    "sha256",
    "error",
]

DEFAULT_XLSX_EXPORT_HEADERS = [
    "saved_at",
    "status",
    "folder",
    "uid",
    "date_header",
    "direction",
    "mittente_vero",
    "header_to",
    "subject_header",
    "filename",
    "relative_path",
    "size_bytes",
    "sha256",
    "error",
]

EXPORT_SHEET_NAME = "indice"
INTERNAL_SHEET_NAME = "_internal"
DEFAULT_FILENAME_TEMPLATE = "{date} - {party} - {subject} __{hash}"
DEFAULT_FILENAME_MAX_LENGTH = 180
FILENAME_TEMPLATE_FIELDS = [
    ("date", "filename_date"),
    ("party", "pec_party"),
    ("subject", "subject_for_filename"),
    ("folder", "folder"),
    ("uid", "uid"),
    ("direction", "direction"),
    ("pec_type", "pec_type"),
    ("message_id", "message_id"),
    ("hash", "sha256_short"),
    ("full_hash", "sha256"),
]

WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}

IMAP_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

BASE_DIR = Path(__file__).resolve().parent
I18N_DIR = BASE_DIR / "i18n"
RESOURCES_DIR = BASE_DIR / "resources"
APP_ICON_PATH = RESOURCES_DIR / "app_icon_256.png"
SUPPORTED_LANGUAGES = {"it": "Italiano", "en": "English"}
CURRENT_LANGUAGE = "it"
TRANSLATIONS: Dict[str, str] = {}


def load_translation_catalog(language: str) -> Dict[str, str]:
    """Load a JSON translation catalog from the local i18n directory."""
    lang = language if language in SUPPORTED_LANGUAGES else "it"
    path = I18N_DIR / f"{lang}.json"
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items()}
    except Exception:
        pass
    return {}


def set_language(language: str) -> None:
    """Set the active language used by the UI and worker messages."""
    global CURRENT_LANGUAGE, TRANSLATIONS
    CURRENT_LANGUAGE = language if language in SUPPORTED_LANGUAGES else "it"
    TRANSLATIONS = load_translation_catalog(CURRENT_LANGUAGE)


def tr(key: str, **kwargs: Any) -> str:
    """Translate a message key and interpolate named placeholders safely."""
    text = TRANSLATIONS.get(key, key)
    if kwargs:
        try:
            return text.format(**kwargs)
        except Exception:
            return text
    return text


set_language(CURRENT_LANGUAGE)


class OperationCancelled(RuntimeError):
    """Cancellation requested by the user through the Stop button."""


def human_size(num_bytes: int) -> str:
    try:
        value = float(num_bytes or 0)
    except Exception:
        value = 0.0
    units = ["B", "KB", "MB", "GB", "TB"]
    idx = 0
    while value >= 1024 and idx < len(units) - 1:
        value /= 1024
        idx += 1
    if idx == 0:
        return f"{int(value)} B"
    return f"{value:.2f} {units[idx]}"


# ---------------------------------------------------------------------------
# Text, MIME and PEC
# ---------------------------------------------------------------------------


def decode_mime_header(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return value


def addresses_to_string(value: Optional[str]) -> str:
    if not value:
        return ""
    decoded = decode_mime_header(value)
    pairs = getaddresses([decoded])
    items = []
    for name, addr in pairs:
        if addr:
            items.append(addr)
        elif name:
            items.append(name)
    return "; ".join(items)


def first_email(value: str) -> str:
    _, addr = parseaddr(value or "")
    return (addr or value or "").strip()


def normalize_email(value: str) -> str:
    return first_email(value).strip().lower()


def parse_email_date(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat()
    except Exception:
        return decode_mime_header(value)


def parse_email_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def local_filename_datetime(msg: Message) -> str:
    dt = parse_email_datetime(msg.get("date"))
    if dt is None:
        return datetime.now().strftime("%Y-%m-%d_%H%M%S")
    try:
        dt = dt.astimezone()
    except Exception:
        pass
    return dt.strftime("%Y-%m-%d_%H%M%S")


def read_part_bytes(part: Message) -> bytes:
    payload = part.get_payload(decode=True)
    if payload is not None:
        return payload
    text = part.get_payload()
    if isinstance(text, str):
        charset = part.get_content_charset() or "utf-8"
        return text.encode(charset, errors="replace")
    return b""


def strip_xml_namespace(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def xml_text(parent: ET.Element, path: Sequence[str]) -> str:
    node: Optional[ET.Element] = parent
    for name in path:
        if node is None:
            return ""
        found = None
        for child in list(node):
            if strip_xml_namespace(child.tag) == name:
                found = child
                break
        node = found
    if node is None or node.text is None:
        return ""
    return node.text.strip()


def xml_all_text(parent: ET.Element, path: Sequence[str]) -> List[str]:
    if not path:
        return []
    node: Optional[ET.Element] = parent
    for name in path[:-1]:
        if node is None:
            return []
        found = None
        for child in list(node):
            if strip_xml_namespace(child.tag) == name:
                found = child
                break
        node = found
    if node is None:
        return []
    last = path[-1]
    out: List[str] = []
    for child in list(node):
        if strip_xml_namespace(child.tag) == last and child.text:
            out.append(child.text.strip())
    return out


def find_daticert_xml(msg: Message) -> Optional[bytes]:
    for part in msg.walk():
        filename = decode_mime_header(part.get_filename() or "").lower()
        ctype = (part.get_content_type() or "").lower()
        if filename == "daticert.xml" or (filename.endswith(".xml") and "daticert" in filename):
            data = read_part_bytes(part)
            if b"<postacert" in data[:500].lower() or b"postacert" in data[:500].lower():
                return data
        if ctype in {"application/xml", "text/xml"}:
            data = read_part_bytes(part)
            if b"postacert" in data[:500].lower():
                return data
    return None


def parse_daticert_xml(data: bytes) -> Dict[str, Any]:
    try:
        root = ET.fromstring(data)
    except Exception:
        return {}
    tipo = root.attrib.get("tipo", "")
    mittente = xml_text(root, ["intestazione", "mittente"])
    destinatari = xml_all_text(root, ["intestazione", "destinatari"])
    risposte = xml_text(root, ["intestazione", "risposte"])
    oggetto = xml_text(root, ["intestazione", "oggetto"])
    consegna = xml_text(root, ["dati", "consegna"])
    gestore = xml_text(root, ["dati", "gestore-emittente"])
    giorno = xml_text(root, ["dati", "data", "giorno"])
    ora = xml_text(root, ["dati", "data", "ora"])
    identificativo = xml_text(root, ["dati", "identificativo"])
    msgid = xml_text(root, ["dati", "msgid"])
    return {
        "tipo": tipo,
        "mittente": mittente,
        "destinatari": destinatari,
        "risposte": risposte,
        "oggetto": oggetto,
        "consegna": consegna,
        "gestore": gestore,
        "giorno": giorno,
        "ora": ora,
        "identificativo": identificativo,
        "msgid": msgid,
    }


def find_first_embedded_rfc822(msg: Message) -> Optional[Message]:
    for part in msg.walk():
        if (part.get_content_type() or "").lower() == "message/rfc822":
            payload = part.get_payload()
            if isinstance(payload, list) and payload:
                candidate = payload[0]
                if isinstance(candidate, Message):
                    return candidate
            raw = read_part_bytes(part)
            if raw:
                try:
                    return email.message_from_bytes(raw, policy=email.policy.default)
                except Exception:
                    return None
    return None


def choose_pec_party(daticert: Dict[str, Any], header_from: str, own_email: str = "") -> str:
    """Choose the most useful counterparty for file names and the index.

    Practical rules:
    - actual PEC message: sender from daticert.xml;
    - receipts/deliveries: delivery recipient when present, otherwise the first
      external recipient, otherwise the sender;
    - fallback: already-decoded From/Reply-To.
    """
    own = normalize_email(own_email)
    tipo = (daticert.get("tipo") or "").strip().lower()
    mittente = daticert.get("mittente") or ""
    consegna = daticert.get("consegna") or ""
    destinatari = daticert.get("destinatari") or []

    if tipo == "posta-certificata" and mittente:
        return mittente

    if consegna:
        return consegna

    for dest in destinatari:
        if not own or normalize_email(dest) != own:
            return dest

    if mittente:
        return mittente

    # Fallback for "Per conto di: x@y" in the display name.
    decoded = decode_mime_header(header_from)
    m = re.search(r"per\s+conto\s+di\s*:\s*([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})", decoded, re.I)
    if m:
        return m.group(1)

    return first_email(decoded)


def infer_direction(header_from: str, header_to: str, daticert: Dict[str, Any], own_email: str, folder: str) -> str:
    folder_l = folder.lower()
    if any(x in folder_l for x in ["sent", "inviat", "outbox", "posta inviata"]):
        return "uscita"

    own = normalize_email(own_email)
    from_addr = normalize_email(header_from)
    to_addrs = {normalize_email(x) for x in re.split(r"[;,]", header_to or "") if x.strip()}

    if own and from_addr == own:
        return "uscita"
    if own and own in to_addrs:
        return "ingresso"

    tipo = (daticert.get("tipo") or "").lower()
    mittente = normalize_email(daticert.get("mittente") or "")
    if tipo and own and mittente == own and tipo != "posta-certificata":
        return "uscita/ricevuta"
    if tipo == "posta-certificata":
        return "ingresso"
    return ""


def parse_full_metadata(raw: bytes, folder: str, uid: str, internal_date: str, own_email: str) -> Dict[str, Any]:
    msg = email.message_from_bytes(raw, policy=email.policy.default)
    header_from = addresses_to_string(msg.get("from"))
    header_to = addresses_to_string(msg.get("to"))
    subject_header = decode_mime_header(msg.get("subject"))
    date_header = parse_email_date(msg.get("date"))
    message_id = decode_mime_header(msg.get("message-id"))

    daticert_data = find_daticert_xml(msg)
    daticert: Dict[str, Any] = parse_daticert_xml(daticert_data) if daticert_data else {}

    embedded = find_first_embedded_rfc822(msg)
    embedded_from = addresses_to_string(embedded.get("from")) if embedded is not None else ""
    embedded_subject = decode_mime_header(embedded.get("subject")) if embedded is not None else ""

    pec_party = choose_pec_party(daticert, msg.get("from", ""), own_email=own_email)
    if not pec_party and embedded_from:
        pec_party = embedded_from
    if not pec_party:
        pec_party = header_from

    pec_oggetto = daticert.get("oggetto") or embedded_subject or subject_header
    subject_for_filename = pec_oggetto or subject_header

    return {
        "folder": folder,
        "uid": uid,
        "message_id": message_id,
        "date_header": date_header,
        "internal_date": internal_date,
        "direction": infer_direction(header_from, header_to, daticert, own_email, folder),
        "mittente_vero": pec_party,
        "header_from": header_from,
        "header_to": header_to,
        "subject_header": subject_header,
        "pec_type": daticert.get("tipo", ""),
        "pec_party": pec_party,
        "pec_mittente": daticert.get("mittente", ""),
        "pec_destinatari": "; ".join(daticert.get("destinatari", []) or []),
        "pec_oggetto": pec_oggetto,
        "subject_for_filename": subject_for_filename,
        "filename_date": local_filename_datetime(msg),
    }


def is_likely_pec_preview(row: Dict[str, Any]) -> bool:
    """Decide whether preview should download RAW data to find daticert.xml.

    The real PEC sender is often not in the visible SMTP From header but inside
    daticert.xml or postacert.eml. Downloading RAW for every message can be slow,
    so this is limited to messages that look like PEC.
    """
    text = " ".join(
        str(row.get(k, ""))
        for k in ("header_from", "header_to", "subject_header", "message_id")
    ).lower()
    pec_tokens = [
        "posta-certificata@",
        "postacert",
        "posta certificata",
        "posta-certificata",
        "accettazione:",
        "consegna:",
        "ricevuta",
        "avvenuta consegna",
        "mancata consegna",
    ]
    return any(token in text for token in pec_tokens)


def parse_header_preview(header_bytes: bytes, folder: str, uid: str, internal_date: str, size: int) -> Dict[str, Any]:
    msg = email.message_from_bytes(header_bytes, policy=email.policy.default)
    header_from = addresses_to_string(msg.get("from"))
    return {
        "folder": folder,
        "uid": uid,
        "date_header": parse_email_date(msg.get("date")),
        "internal_date": internal_date,
        "header_from": header_from,
        "mittente_vero": header_from,
        "display_from": header_from,
        "header_to": addresses_to_string(msg.get("to")),
        "subject_header": decode_mime_header(msg.get("subject")),
        "message_id": decode_mime_header(msg.get("message-id")),
        "pec_party": "",
        "pec_type": "",
        "pec_mittente": "",
        "pec_destinatari": "",
        "pec_oggetto": "",
        "size_bytes": size,
        "preview_mode": "header",
    }


# ---------------------------------------------------------------------------
# File names safe for Windows/Linux/macOS
# ---------------------------------------------------------------------------


def sanitize_filesystem_text(value: str, fallback: str = "senza_nome") -> str:
    """Return text that is safe as one file-system path component.

    The output avoids Windows/NTFS forbidden characters, ASCII control
    characters, trailing dots/spaces and reserved DOS device names. The same
    constraints are also safe on Linux and macOS.
    """
    value = unicodedata.normalize("NFKC", value or "")
    value = value.replace("\x00", "")
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.strip(" .")
    if not value:
        value = fallback
    if value.upper() in WINDOWS_RESERVED_NAMES:
        value = f"_{value}"
    return value or fallback


def safe_component(value: str, max_len: int = 90, fallback: str = "senza_nome") -> str:
    value = sanitize_filesystem_text(value, fallback=fallback)
    if len(value) > max_len:
        value = value[:max_len].rstrip(" ._")
    return value or fallback


def clamp_filename_max_length(value: Any) -> int:
    try:
        max_len = int(value)
    except Exception:
        max_len = DEFAULT_FILENAME_MAX_LENGTH
    return min(max(max_len, 40), 240)


def safe_eml_filename(base: str, sha256_hex: str, max_len: int) -> str:
    """Build a filesystem-safe .eml filename whose total length is <= max_len."""
    max_len = clamp_filename_max_length(max_len)
    short_hash = (sha256_hex or "")[:12] or "nohash"
    suffix = f" __{short_hash}"
    clean = sanitize_filesystem_text(base, fallback="message")
    if not clean.lower().endswith(short_hash.lower()):
        clean = f"{clean}{suffix}"
    ext = ".eml"
    max_base_len = max_len - len(ext)
    if len(clean) > max_base_len:
        if clean.lower().endswith(short_hash.lower()):
            # Preserve the short hash at the end, because it prevents collisions
            # when long subjects are truncated.
            allowed_prefix = max_base_len - len(suffix)
            if allowed_prefix < 1:
                clean = short_hash[:max_base_len]
            else:
                clean = clean[:allowed_prefix].rstrip(" ._") + suffix
        else:
            clean = clean[:max_base_len].rstrip(" ._")
    clean = sanitize_filesystem_text(clean, fallback="message")
    if len(clean) > max_base_len:
        clean = clean[:max_base_len].rstrip(" ._")
    return f"{clean or 'message'}{ext}"


def filename_template_values(meta: Dict[str, Any], sha256_hex: str) -> Dict[str, str]:
    return {
        "date": str(meta.get("filename_date") or "senza_data"),
        "party": str(meta.get("pec_party") or meta.get("mittente_vero") or meta.get("header_from") or "sconosciuto"),
        "subject": str(meta.get("subject_for_filename") or meta.get("subject_header") or "senza_oggetto"),
        "folder": str(meta.get("folder") or "cartella"),
        "uid": str(meta.get("uid") or "uid"),
        "direction": str(meta.get("direction") or ""),
        "pec_type": str(meta.get("pec_type") or ""),
        "message_id": str(meta.get("message_id") or ""),
        "hash": (sha256_hex or "")[:12],
        "full_hash": sha256_hex or "",
    }


def build_filename(
    meta: Dict[str, Any],
    sha256_hex: str,
    template: str = DEFAULT_FILENAME_TEMPLATE,
    max_len: int = DEFAULT_FILENAME_MAX_LENGTH,
) -> str:
    values = filename_template_values(meta, sha256_hex)
    template = (template or DEFAULT_FILENAME_TEMPLATE).strip() or DEFAULT_FILENAME_TEMPLATE
    try:
        base = template.format(**values)
    except Exception:
        base = DEFAULT_FILENAME_TEMPLATE.format(**values)
    return safe_eml_filename(base, sha256_hex, max_len)


def with_filename_suffix(path: Path, suffix: str, max_len: int) -> Path:
    max_len = clamp_filename_max_length(max_len)
    ext = path.suffix or ".eml"
    max_stem_len = max(max_len - len(ext) - len(suffix), 1)
    stem = sanitize_filesystem_text(path.stem, fallback="message")
    new_stem = stem[:max_stem_len].rstrip(" ._") + suffix
    return path.with_name(f"{new_stem}{ext}")


def atomic_write_bytes(target: Path, data: bytes) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=target.name + ".", suffix=".tmp", dir=str(target.parent))
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_name, target)
    except Exception:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# XLSX index
# ---------------------------------------------------------------------------


def normalize_export_headers(headers: Optional[Sequence[str]]) -> List[str]:
    selected: List[str] = []
    for header in headers or DEFAULT_XLSX_EXPORT_HEADERS:
        header = str(header).strip()
        if header in INDEX_HEADERS and header not in selected:
            selected.append(header)
    if not selected:
        selected = list(DEFAULT_XLSX_EXPORT_HEADERS)
    return selected


class IndexManager:
    def __init__(self, index_path: Path, export_headers: Optional[Sequence[str]] = None):
        self.index_path = index_path
        self.export_headers = normalize_export_headers(export_headers)
        self.wb: Workbook
        self.ws: Worksheet
        self.header_to_col: Dict[str, int] = {}
        self.hash_to_row: Dict[str, int] = {}
        self._load()

    def _load(self) -> None:
        if self.index_path.exists():
            self.wb = load_workbook(self.index_path)
            if INTERNAL_SHEET_NAME in self.wb.sheetnames:
                self.ws = self.wb[INTERNAL_SHEET_NAME]
            else:
                self.ws = self.wb.active
                # Existing indexes used a single visible sheet named "indice".
                # Keep its data and move it to the hidden internal sheet so the
                # visible sheet can be regenerated with the selected columns.
                self.ws.title = INTERNAL_SHEET_NAME
            if self.ws.max_row < 1:
                self.ws.append(INDEX_HEADERS)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = INTERNAL_SHEET_NAME
            self.ws.append(INDEX_HEADERS)
        self._ensure_headers()
        self._backfill_derived_columns()
        self._rebuild_maps()

    def _ensure_headers(self) -> None:
        current = [self.ws.cell(1, c).value for c in range(1, self.ws.max_column + 1)]
        for h in INDEX_HEADERS:
            if h not in current:
                self.ws.cell(1, self.ws.max_column + 1, h)
                current.append(h)
        self.header_to_col = {str(self.ws.cell(1, c).value): c for c in range(1, self.ws.max_column + 1)}

    def _backfill_derived_columns(self) -> None:
        """Populate columns introduced by newer index versions.

        In earlier versions, the value shown in the GUI as "real sender" was
        stored only in the technical pec_party column. It is now duplicated into
        mittente_vero so the XLSX is easier to filter.
        """
        mv_col = self.header_to_col.get("mittente_vero")
        if not mv_col:
            return
        party_col = self.header_to_col.get("pec_party")
        from_col = self.header_to_col.get("header_from")
        for r in range(2, self.ws.max_row + 1):
            if self.ws.cell(r, mv_col).value:
                continue
            value = ""
            if party_col:
                value = self.ws.cell(r, party_col).value or ""
            if not value and from_col:
                value = self.ws.cell(r, from_col).value or ""
            if value:
                self.ws.cell(r, mv_col, value)

    def _rebuild_maps(self) -> None:
        self.hash_to_row.clear()
        col = self.header_to_col.get("sha256")
        if not col:
            return
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, col).value
            if val:
                self.hash_to_row[str(val)] = r

    def has_hash(self, sha256_hex: str) -> bool:
        return sha256_hex in self.hash_to_row

    def existing_row_path(self, sha256_hex: str, base_dir: Path) -> Optional[Path]:
        row = self.hash_to_row.get(sha256_hex)
        if not row:
            return None
        col = self.header_to_col.get("relative_path")
        if not col:
            return None
        rel = self.ws.cell(row, col).value
        if not rel:
            return None
        return base_dir / str(rel)

    def append_or_update(self, record: Dict[str, Any]) -> None:
        sha = record.get("sha256")
        row = self.hash_to_row.get(sha) if sha else None
        if row is None:
            row = self.ws.max_row + 1
            if sha:
                self.hash_to_row[sha] = row
        for key, value in record.items():
            if key not in self.header_to_col:
                self.ws.cell(1, self.ws.max_column + 1, key)
                self._ensure_headers()
            self.ws.cell(row, self.header_to_col[key], value)

    def mark_deleted(self, sha256_hex: str, deleted_at: str) -> None:
        row = self.hash_to_row.get(sha256_hex)
        if not row:
            return
        self.ws.cell(row, self.header_to_col["status"], "deleted")
        self.ws.cell(row, self.header_to_col["deleted_at"], deleted_at)

    def _refresh_export_sheet(self, export_headers: Optional[Sequence[str]] = None) -> None:
        headers = normalize_export_headers(export_headers or self.export_headers)
        if EXPORT_SHEET_NAME in self.wb.sheetnames and self.wb[EXPORT_SHEET_NAME] is not self.ws:
            self.wb.remove(self.wb[EXPORT_SHEET_NAME])
        if self.ws.title == EXPORT_SHEET_NAME:
            self.ws.title = INTERNAL_SHEET_NAME
        export_ws = self.wb.create_sheet(EXPORT_SHEET_NAME, 0)
        export_ws.append(headers)
        for r in range(2, self.ws.max_row + 1):
            export_ws.append([self.ws.cell(r, self.header_to_col[h]).value for h in headers])
        export_ws.freeze_panes = "A2"
        if export_ws.max_row >= 1 and export_ws.max_column >= 1:
            export_ws.auto_filter.ref = export_ws.dimensions
        for c, header in enumerate(headers, start=1):
            max_width = len(header)
            for r in range(2, min(export_ws.max_row, 200) + 1):
                value = export_ws.cell(r, c).value
                if value is not None:
                    max_width = max(max_width, min(len(str(value)), 60))
            export_ws.column_dimensions[export_ws.cell(1, c).column_letter].width = min(max_width + 2, 64)
        self.ws.sheet_state = "hidden"
        self.wb.active = self.wb.index(export_ws)

    def save(self, export_headers: Optional[Sequence[str]] = None) -> None:
        self._refresh_export_sheet(export_headers)
        self.index_path.parent.mkdir(parents=True, exist_ok=True)
        fd, tmp_name = tempfile.mkstemp(prefix=self.index_path.name + ".", suffix=".xlsx", dir=str(self.index_path.parent))
        os.close(fd)
        try:
            self.wb.save(tmp_name)
            os.replace(tmp_name, self.index_path)
        except Exception:
            try:
                os.unlink(tmp_name)
            except OSError:
                pass
            raise


# ---------------------------------------------------------------------------
# IMAP helpers, including modified UTF-7 for folder names
# ---------------------------------------------------------------------------


def encode_imap_utf7(text: str) -> bytes:
    out = bytearray()
    buf = bytearray()

    def flush_buf() -> None:
        nonlocal buf
        if not buf:
            return
        b64 = base64.b64encode(bytes(buf)).decode("ascii").rstrip("=").replace("/", ",")
        out.extend(b"&" + b64.encode("ascii") + b"-")
        buf = bytearray()

    for ch in text:
        code = ord(ch)
        if 0x20 <= code <= 0x7E:
            flush_buf()
            if ch == "&":
                out.extend(b"&-")
            else:
                out.append(code)
        else:
            buf.extend(ch.encode("utf-16-be"))
    flush_buf()
    return bytes(out)


def decode_imap_utf7(data: bytes) -> str:
    text = data.decode("ascii", errors="replace")
    out: List[str] = []
    i = 0
    while i < len(text):
        if text[i] != "&":
            out.append(text[i])
            i += 1
            continue
        j = text.find("-", i)
        if j == -1:
            out.append("&")
            i += 1
            continue
        token = text[i + 1 : j]
        if token == "":
            out.append("&")
        else:
            padded = token.replace(",", "/")
            padded += "=" * ((4 - len(padded) % 4) % 4)
            try:
                out.append(base64.b64decode(padded).decode("utf-16-be"))
            except Exception:
                out.append("&" + token + "-")
        i = j + 1
    return "".join(out)


def quote_mailbox(folder: str) -> bytes:
    encoded = encode_imap_utf7(folder)
    escaped = encoded.replace(b"\\", b"\\\\").replace(b'"', b'\\"')
    return b'"' + escaped + b'"'


def imap_before_date(dt: datetime) -> str:
    return f"{dt.day:02d}-{IMAP_MONTHS[dt.month - 1]}-{dt.year:04d}"


def parse_fetch_meta(fetch_prefix: bytes) -> Tuple[str, int]:
    text = fetch_prefix.decode("latin-1", errors="replace")
    internal = ""
    size = 0
    m = re.search(r'INTERNALDATE\s+"([^"]+)"', text, re.I)
    if m:
        internal = m.group(1)
    m = re.search(r"RFC822\.SIZE\s+(\d+)", text, re.I)
    if m:
        size = int(m.group(1))
    return internal, size


def parse_list_line(line: bytes) -> str:
    # Typical lines: b'(\\HasNoChildren) "/" "INBOX"'
    m = re.search(br'"((?:[^"\\]|\\.)*)"\s*$', line)
    if m:
        raw = m.group(1).replace(b'\\"', b'"').replace(b"\\\\", b"\\")
        return decode_imap_utf7(raw)
    parts = line.split()
    if parts:
        return decode_imap_utf7(parts[-1].strip(b'"'))
    return line.decode("latin-1", errors="replace")


class ImapConnection:
    def __init__(
        self,
        host: str,
        port: int,
        ssl_mode: str,
        username: str,
        password: str,
        timeout: int = 60,
        debug_enabled: bool = False,
        debug_callback=None,
    ):
        self.host = host
        self.port = port
        self.ssl_mode = ssl_mode
        self.username = username
        self.password = password
        self.timeout = timeout
        self.debug_enabled = debug_enabled
        self.debug_callback = debug_callback
        self.conn: Optional[imaplib.IMAP4] = None

    def debug(self, text: str) -> None:
        if self.debug_enabled and self.debug_callback is not None:
            self.debug_callback("DEBUG IMAP: " + text)

    @staticmethod
    def _short_data(data: Any, max_len: int = 1200) -> str:
        text = repr(data)
        if len(text) > max_len:
            return text[:max_len] + tr("debug_truncated")
        return text

    def __enter__(self) -> "ImapConnection":
        self.connect()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def connect(self) -> None:
        self.debug(tr("debug_open_connection", host=self.host, port=self.port, ssl=self.ssl_mode, timeout=self.timeout))
        context = ssl.create_default_context()
        try:
            if self.ssl_mode == "SSL/TLS":
                self.debug(tr("debug_create_ssl_socket"))
                self.conn = imaplib.IMAP4_SSL(self.host, self.port, ssl_context=context, timeout=self.timeout)
            else:
                self.debug(tr("debug_create_plain_socket"))
                self.conn = imaplib.IMAP4(self.host, self.port, timeout=self.timeout)
                if self.ssl_mode == "STARTTLS":
                    self.debug(tr("debug_request_starttls"))
                    typ, data = self.conn.starttls(ssl_context=context)
                    self.debug(f"STARTTLS -> {typ} {self._short_data(data)}")
                    if typ != "OK":
                        raise RuntimeError(tr("error_starttls_failed", data=data))
        except Exception as e:
            raise RuntimeError(tr("error_imap_connection_failed", host=self.host, port=self.port, ssl=self.ssl_mode, error=e)) from e

        self.debug(tr("debug_socket_open", welcome=self._short_data(getattr(self.conn, 'welcome', b''))))
        try:
            typ, data = self.conn.capability()
            self.debug(f"CAPABILITY pre-login -> {typ} {self._short_data(data)}")
        except Exception as e:
            self.debug(tr("debug_capability_pre_unavailable", error=e))

        self.debug(tr("debug_login_user", username=repr(self.username)))
        try:
            typ, data = self.conn.login(self.username, self.password)
        except imaplib.IMAP4.error as e:
            raise RuntimeError(tr("error_login_rejected", error=e)) from e
        except Exception as e:
            raise RuntimeError(tr("error_login_failed", error=e)) from e
        self.debug(f"LOGIN -> {typ} {self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_login_failed", error=data))
        try:
            typ, data = self.conn.capability()
            self.debug(f"CAPABILITY post-login -> {typ} {self._short_data(data)}")
        except Exception as e:
            self.debug(tr("debug_capability_post_unavailable", error=e))

    def abort(self) -> None:
        """Try to unblock any pending IMAP calls immediately.

        imaplib does not provide real asynchronous cancellation for FETCH/SEARCH/LIST;
        closing the socket makes the current call fail and lets the worker finish
        without waiting for the full timeout.
        """
        conn = self.conn
        if conn is None:
            return
        self.debug(tr("debug_abort_socket"))
        sock = getattr(conn, "sock", None)
        if sock is not None:
            try:
                sock.shutdown(socket.SHUT_RDWR)
            except Exception as e:
                self.debug(tr("debug_socket_shutdown_ignored", error=e))
            try:
                sock.close()
            except Exception as e:
                self.debug(tr("debug_socket_close_ignored", error=e))
        try:
            conn.shutdown()
        except Exception as e:
            self.debug(tr("debug_imap_shutdown_ignored", error=e))

    def close(self) -> None:
        if self.conn is None:
            return
        try:
            # IMPORTANT: do not use IMAP CLOSE here. In IMAP, CLOSE permanently removes
            # messages marked \Deleted in the selected mailbox,
            # which is effectively an implicit expunge. We want permanent deletion
            # to happen only when the user explicitly enables
            # EXPUNGE. LOGOUT closes the session without sending CLOSE.
            self.debug(tr("debug_logout_without_close"))
            try:
                self.conn.logout()
                self.debug(tr("debug_logout_done"))
            except Exception as e:
                self.debug(tr("debug_logout_ignored", error=e))
        finally:
            self.conn = None

    def require(self) -> imaplib.IMAP4:
        if self.conn is None:
            raise RuntimeError(tr("error_imap_not_open"))
        return self.conn

    def list_folders(self) -> List[str]:
        conn = self.require()
        self.debug(tr("debug_list_command"))
        typ, data = conn.list()
        self.debug(f"LIST -> {typ}; righe={len(data or [])}; dati={self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_list_failed", data=data))
        folders = []
        for line in data or []:
            if isinstance(line, bytes):
                folders.append(parse_list_line(line))
        return sorted(set(folders), key=str.lower)

    def select_folder(self, folder: str) -> None:
        conn = self.require()
        quoted = quote_mailbox(folder)
        self.debug(tr("debug_select_folder", folder=repr(folder), encoded=repr(quoted)))
        typ, data = conn.select(quoted, readonly=False)
        self.debug(f"SELECT {folder!r} -> {typ} {self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_select_folder_failed", folder=folder, data=data))

    def search_filtered(self, cutoff: datetime, min_size_bytes: int = 0) -> List[str]:
        conn = self.require()
        before = imap_before_date(cutoff)
        criteria: List[Any] = ["BEFORE", before]
        if min_size_bytes and min_size_bytes > 0:
            # IMAP LARGER is strictly "> N octets", matching the requested filter.
            criteria.extend(["LARGER", str(int(min_size_bytes))])
        self.debug(tr("debug_search_criteria", criteria=repr(criteria)))
        typ, data = conn.uid("SEARCH", None, *criteria)
        self.debug(f"SEARCH -> {typ}; dati={self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_search_failed", data=data))
        if not data or not data[0]:
            return []
        uids = data[0].decode("ascii", errors="ignore").split()
        self.debug(tr("debug_search_found", count=len(uids)))
        return uids

    # Internal compatibility with previous script versions.
    def search_before(self, cutoff: datetime) -> List[str]:
        return self.search_filtered(cutoff, 0)

    def fetch_header(self, uid: str) -> Tuple[bytes, str, int]:
        conn = self.require()
        self.debug(f"FETCH HEADER UID {uid}…")
        typ, data = conn.uid("FETCH", uid, "(BODY.PEEK[HEADER] RFC822.SIZE INTERNALDATE)")
        self.debug(f"FETCH HEADER UID {uid} -> {typ}; parti={len(data or [])}")
        if typ != "OK":
            raise RuntimeError(tr("error_fetch_header_failed", uid=uid, data=data))
        header = b""
        internal = ""
        size = 0
        for item in data or []:
            if isinstance(item, tuple):
                prefix, payload = item
                internal, size = parse_fetch_meta(prefix)
                header = payload or b""
                break
        self.debug(f"FETCH HEADER UID {uid}: header_bytes={len(header)}, size={size}, internal_date={internal!r}")
        return header, internal, size

    def fetch_full_raw(self, uid: str) -> Tuple[bytes, str, int]:
        conn = self.require()
        self.debug(f"FETCH FULL RAW UID {uid}…")
        typ, data = conn.uid("FETCH", uid, "(BODY.PEEK[] RFC822.SIZE INTERNALDATE)")
        self.debug(f"FETCH FULL RAW UID {uid} -> {typ}; parti={len(data or [])}")
        if typ != "OK":
            raise RuntimeError(tr("error_fetch_full_failed", uid=uid, data=data))
        raw = b""
        internal = ""
        size = 0
        for item in data or []:
            if isinstance(item, tuple):
                prefix, payload = item
                internal, size = parse_fetch_meta(prefix)
                raw = payload or b""
                break
        if not raw:
            raise RuntimeError(tr("error_empty_message", uid=uid))
        self.debug(f"FETCH FULL RAW UID {uid}: raw_bytes={len(raw)}, size={size}, internal_date={internal!r}")
        return raw, internal, size

    def mark_deleted(self, uid: str) -> None:
        conn = self.require()
        self.debug(f"STORE +FLAGS.SILENT (\\Deleted) UID {uid}…")
        typ, data = conn.uid("STORE", uid, "+FLAGS.SILENT", r"(\Deleted)")
        self.debug(f"STORE UID {uid} -> {typ} {self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_store_deleted_failed", uid=uid, data=data))

    def expunge(self) -> None:
        conn = self.require()
        self.debug("EXPUNGE…")
        typ, data = conn.expunge()
        self.debug(f"EXPUNGE -> {typ} {self._short_data(data)}")
        if typ != "OK":
            raise RuntimeError(tr("error_expunge_failed", data=data))

# ---------------------------------------------------------------------------
# Worker PySide
# ---------------------------------------------------------------------------


@dataclass
class RuntimeConfig:
    host: str
    port: int
    ssl_mode: str
    username: str
    password: str
    folders: List[str]
    cutoff: datetime
    min_size_bytes: int
    backup_dir: Path
    ask_delete: bool
    expunge: bool
    timeout_seconds: int = 60
    debug: bool = False
    preview_deep_pec: bool = True
    xlsx_columns: Optional[List[str]] = None
    filename_template: str = DEFAULT_FILENAME_TEMPLATE
    filename_max_length: int = DEFAULT_FILENAME_MAX_LENGTH


class WorkerBase(QObject):
    log = Signal(str)
    progress = Signal(int, int)
    failed = Signal(str)
    cancelled = Signal(str)

    def __init__(self) -> None:
        super().__init__()
        self._cancel_requested = threading.Event()
        self.current_imap: Optional[ImapConnection] = None

    @Slot()
    def request_cancel(self) -> None:
        if self._cancel_requested.is_set():
            return
        self._cancel_requested.set()
        self.log.emit(tr("log_stop_worker_requested"))
        imap = self.current_imap
        if imap is not None:
            try:
                imap.abort()
            except Exception as e:
                self.log.emit(tr("log_stop_abort_failed", error=e))

    def is_cancelled(self) -> bool:
        return self._cancel_requested.is_set()

    def check_cancelled(self) -> None:
        if self.is_cancelled():
            raise OperationCancelled(tr("operation_cancelled_short"))

    def emit_log(self, text: str) -> None:
        self.log.emit(text)

    def emit_failure(self, exc: BaseException) -> None:
        if isinstance(exc, OperationCancelled) or self.is_cancelled():
            self.cancelled.emit(
                tr("operation_cancelled_long")
            )
            return
        cfg = getattr(self, "cfg", None)
        if getattr(cfg, "debug", False):
            tb = traceback.format_exc()
            self.log.emit(tr("technical_traceback") + "\n" + tb)
            self.failed.emit(f"{exc}\n\n{tr('technical_traceback')}:\n{tb}")
        else:
            self.failed.emit(str(exc))


class ListFoldersWorker(WorkerBase):
    done = Signal(list)

    def __init__(self, cfg: RuntimeConfig):
        super().__init__()
        self.cfg = cfg

    @Slot()
    def run(self) -> None:
        try:
            with ImapConnection(
                self.cfg.host,
                self.cfg.port,
                self.cfg.ssl_mode,
                self.cfg.username,
                self.cfg.password,
                timeout=self.cfg.timeout_seconds,
                debug_enabled=self.cfg.debug,
                debug_callback=self.emit_log,
            ) as imap:
                self.current_imap = imap
                self.check_cancelled()
                folders = imap.list_folders()
                self.check_cancelled()
            self.done.emit(folders)
        except Exception as e:
            self.emit_failure(e)
        finally:
            self.current_imap = None


class PreviewWorker(WorkerBase):
    done = Signal(list)

    def __init__(self, cfg: RuntimeConfig):
        super().__init__()
        self.cfg = cfg

    @Slot()
    def run(self) -> None:
        rows: List[Dict[str, Any]] = []
        try:
            with ImapConnection(
                self.cfg.host,
                self.cfg.port,
                self.cfg.ssl_mode,
                self.cfg.username,
                self.cfg.password,
                timeout=self.cfg.timeout_seconds,
                debug_enabled=self.cfg.debug,
                debug_callback=self.emit_log,
            ) as imap:
                self.current_imap = imap
                self.check_cancelled()
                total = 0
                folder_uids: Dict[str, List[str]] = {}
                for folder in self.cfg.folders:
                    self.check_cancelled()
                    imap.select_folder(folder)
                    uids = imap.search_filtered(self.cfg.cutoff, self.cfg.min_size_bytes)
                    folder_uids[folder] = uids
                    total += len(uids)
                    self.emit_log(tr("log_folder_preview_count", folder=folder, count=len(uids), cutoff=self.cfg.cutoff.date(), size_suffix=(tr("log_size_suffix", size=human_size(self.cfg.min_size_bytes)) if self.cfg.min_size_bytes else ".")))
                done = 0
                for folder, uids in folder_uids.items():
                    self.check_cancelled()
                    imap.select_folder(folder)
                    for uid in uids:
                        self.check_cancelled()
                        header, internal, size = imap.fetch_header(uid)
                        row = parse_header_preview(header, folder, uid, internal, size)
                        if self.cfg.preview_deep_pec and is_likely_pec_preview(row):
                            # For PEC, the SMTP From header is often the provider (posta-certificata@...).
                            # Download RAW only for these cases to read daticert.xml / postacert.eml
                            # and show the real sender/counterparty in the table.
                            raw, internal_full, size_full = imap.fetch_full_raw(uid)
                            meta = parse_full_metadata(raw, folder, uid, internal_full or internal, self.cfg.username)
                            row.update(meta)
                            row["display_from"] = meta.get("pec_party") or meta.get("header_from") or row.get("header_from", "")
                            row["mittente_vero"] = row["display_from"]
                            row["size_bytes"] = len(raw) if raw else (size_full or size)
                            row["internal_date"] = internal_full or internal
                            row["preview_mode"] = "pec_profonda"
                        rows.append(row)
                        done += 1
                        self.progress.emit(done, max(total, 1))
            self.check_cancelled()
            self.done.emit(rows)
        except Exception as e:
            self.emit_failure(e)
        finally:
            self.current_imap = None


class BackupWorker(WorkerBase):
    done = Signal(dict)

    def __init__(self, cfg: RuntimeConfig):
        super().__init__()
        self.cfg = cfg

    @Slot()
    def run(self) -> None:
        backup_dir = self.cfg.backup_dir
        index = IndexManager(backup_dir / INDEX_FILENAME, self.cfg.xlsx_columns)
        candidates: List[Dict[str, str]] = []
        saved_count = 0
        skipped_count = 0
        failed_count = 0
        failures: List[str] = []
        total = 0
        folder_uids: Dict[str, List[str]] = {}

        try:
            with ImapConnection(
                self.cfg.host,
                self.cfg.port,
                self.cfg.ssl_mode,
                self.cfg.username,
                self.cfg.password,
                timeout=self.cfg.timeout_seconds,
                debug_enabled=self.cfg.debug,
                debug_callback=self.emit_log,
            ) as imap:
                self.current_imap = imap
                self.check_cancelled()
                for folder in self.cfg.folders:
                    self.check_cancelled()
                    imap.select_folder(folder)
                    uids = imap.search_filtered(self.cfg.cutoff, self.cfg.min_size_bytes)
                    folder_uids[folder] = uids
                    total += len(uids)
                    self.emit_log(tr("log_folder_backup_count", folder=folder, count=len(uids), size_suffix=(tr("log_size_suffix_simple", size=human_size(self.cfg.min_size_bytes)) if self.cfg.min_size_bytes else ".")))

                done = 0
                for folder, uids in folder_uids.items():
                    self.check_cancelled()
                    imap.select_folder(folder)
                    for uid in uids:
                        self.check_cancelled()
                        done += 1
                        self.progress.emit(done, max(total, 1))
                        try:
                            raw, internal, size_from_imap = imap.fetch_full_raw(uid)
                            sha = hashlib.sha256(raw).hexdigest()
                            meta = parse_full_metadata(raw, folder, uid, internal, self.cfg.username)
                            filename = build_filename(meta, sha, self.cfg.filename_template, self.cfg.filename_max_length)
                            folder_safe = safe_component(folder.replace("/", "_"), 80, "cartella")
                            target_dir = backup_dir / folder_safe
                            target_path = target_dir / filename
                            rel_path = target_path.relative_to(backup_dir)

                            existing_path = None
                            if index.has_hash(sha):
                                existing_path = index.existing_row_path(sha, backup_dir)
                            if existing_path and existing_path.exists():
                                if file_sha256(existing_path) != sha:
                                    raise RuntimeError(f"hash in indice ma file diverso: {existing_path}")
                                target_path = existing_path
                                rel_path = target_path.relative_to(backup_dir)
                                skipped_count += 1
                                status = "already_saved"
                            elif target_path.exists():
                                if file_sha256(target_path) == sha:
                                    skipped_count += 1
                                    status = "already_saved"
                                else:
                                    # Unlikely collision: change the suffix while keeping the hash.
                                    target_path = with_filename_suffix(target_path, "_dup", self.cfg.filename_max_length)
                                    atomic_write_bytes(target_path, raw)
                                    if file_sha256(target_path) != sha:
                                        raise RuntimeError(tr("error_hash_verify_failed"))
                                    saved_count += 1
                                    status = "saved"
                            else:
                                atomic_write_bytes(target_path, raw)
                                if target_path.stat().st_size != len(raw):
                                    raise RuntimeError(tr("error_size_verify_failed"))
                                if file_sha256(target_path) != sha:
                                    raise RuntimeError(tr("error_hash_verify_failed"))
                                saved_count += 1
                                status = "saved"

                            rel_path = target_path.relative_to(backup_dir)
                            record = {
                                "saved_at": datetime.now().isoformat(timespec="seconds"),
                                "status": status,
                                "deleted_at": "",
                                "folder": folder,
                                "uid": uid,
                                "message_id": meta.get("message_id", ""),
                                "date_header": meta.get("date_header", ""),
                                "internal_date": internal,
                                "direction": meta.get("direction", ""),
                                "mittente_vero": meta.get("mittente_vero") or meta.get("pec_party") or meta.get("header_from", ""),
                                "header_from": meta.get("header_from", ""),
                                "header_to": meta.get("header_to", ""),
                                "subject_header": meta.get("subject_header", ""),
                                "pec_type": meta.get("pec_type", ""),
                                "pec_party": meta.get("pec_party", ""),
                                "pec_mittente": meta.get("pec_mittente", ""),
                                "pec_destinatari": meta.get("pec_destinatari", ""),
                                "pec_oggetto": meta.get("pec_oggetto", ""),
                                "filename": target_path.name,
                                "relative_path": str(rel_path),
                                "size_bytes": len(raw) if raw else size_from_imap,
                                "sha256": sha,
                                "error": "",
                            }
                            index.append_or_update(record)
                            index.save(self.cfg.xlsx_columns)  # Save often: useful if the script is interrupted.
                            candidates.append({"folder": folder, "uid": uid, "sha256": sha})
                            self.emit_log(tr("log_uid_saved", uid=uid, folder=folder, filename=target_path.name))
                        except Exception as msg_error:
                            if self.is_cancelled() or isinstance(msg_error, OperationCancelled):
                                raise OperationCancelled(tr("operation_cancelled_short")) from msg_error
                            failed_count += 1
                            text = tr("log_uid_error", uid=uid, folder=folder, error=msg_error)
                            failures.append(text)
                            self.emit_log(text)

            self.check_cancelled()
            result = {
                "saved": saved_count,
                "skipped": skipped_count,
                "failed": failed_count,
                "failures": failures,
                "candidates": candidates,
                "index_path": str(backup_dir / INDEX_FILENAME),
                "backup_dir": str(backup_dir),
                "total": total,
            }
            self.done.emit(result)
        except Exception as e:
            self.emit_failure(e)
        finally:
            self.current_imap = None


class DeleteWorker(WorkerBase):
    done = Signal(dict)

    def __init__(self, cfg: RuntimeConfig, candidates: List[Dict[str, str]], index_path: str):
        super().__init__()
        self.cfg = cfg
        self.candidates = candidates
        self.index_path = index_path

    @Slot()
    def run(self) -> None:
        deleted = 0
        failed = 0
        failures: List[str] = []
        by_folder: Dict[str, List[Dict[str, str]]] = defaultdict(list)
        for c in self.candidates:
            by_folder[c["folder"]].append(c)
        try:
            with ImapConnection(
                self.cfg.host,
                self.cfg.port,
                self.cfg.ssl_mode,
                self.cfg.username,
                self.cfg.password,
                timeout=self.cfg.timeout_seconds,
                debug_enabled=self.cfg.debug,
                debug_callback=self.emit_log,
            ) as imap:
                self.current_imap = imap
                self.check_cancelled()
                total = len(self.candidates)
                done = 0
                for folder, items in by_folder.items():
                    self.check_cancelled()
                    imap.select_folder(folder)
                    for c in items:
                        self.check_cancelled()
                        done += 1
                        try:
                            imap.mark_deleted(c["uid"])
                            deleted += 1
                            self.emit_log(tr("log_deleted_set", uid=c["uid"], folder=folder))
                        except Exception as e:
                            if self.is_cancelled() or isinstance(e, OperationCancelled):
                                raise OperationCancelled(tr("operation_cancelled_short")) from e
                            failed += 1
                            failures.append(tr("log_uid_folder_error", uid=c["uid"], folder=folder, error=e))
                        self.progress.emit(done, max(total, 1))
                    if self.cfg.expunge:
                        self.check_cancelled()
                        imap.expunge()
                        self.emit_log(tr("log_expunge_done", folder=folder))

            self.check_cancelled()
            try:
                index = IndexManager(Path(self.index_path), self.cfg.xlsx_columns)
                deleted_at = datetime.now().isoformat(timespec="seconds")
                for c in self.candidates:
                    index.mark_deleted(c["sha256"], deleted_at)
                index.save(self.cfg.xlsx_columns)
            except Exception as e:
                failures.append(tr("error_index_update_after_delete_failed", error=e))
                failed += 1

            self.check_cancelled()
            self.done.emit({"deleted": deleted, "failed": failed, "failures": failures})
        except Exception as e:
            self.emit_failure(e)
        finally:
            self.current_imap = None


class MailSourceWorker(WorkerBase):
    done = Signal(dict)

    def __init__(self, cfg: RuntimeConfig, folder: str, uid: str, max_preview_bytes: int = 2 * 1024 * 1024):
        super().__init__()
        self.cfg = cfg
        self.folder = folder
        self.uid = uid
        self.max_preview_bytes = max_preview_bytes

    @Slot()
    def run(self) -> None:
        try:
            with ImapConnection(
                self.cfg.host,
                self.cfg.port,
                self.cfg.ssl_mode,
                self.cfg.username,
                self.cfg.password,
                timeout=self.cfg.timeout_seconds,
                debug_enabled=self.cfg.debug,
                debug_callback=self.emit_log,
            ) as imap:
                self.current_imap = imap
                self.check_cancelled()
                imap.select_folder(self.folder)
                self.check_cancelled()
                raw, internal, _size = imap.fetch_full_raw(self.uid)
                self.check_cancelled()

            meta = parse_full_metadata(raw, self.folder, self.uid, internal, self.cfg.username)
            truncated = len(raw) > self.max_preview_bytes
            raw_for_text = raw[: self.max_preview_bytes] if truncated else raw
            source = raw_for_text.decode("utf-8", errors="replace")
            if truncated:
                source += "\n\n" + tr("raw_preview_truncated", full_size=human_size(len(raw)), shown_size=human_size(self.max_preview_bytes))
            self.done.emit(
                {
                    "folder": self.folder,
                    "uid": self.uid,
                    "source": source,
                    "truncated": truncated,
                    "raw_bytes": len(raw),
                    "internal_date": internal,
                    "meta": meta,
                }
            )
        except Exception as e:
            self.emit_failure(e)
        finally:
            self.current_imap = None


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class AdvancedConfigDialog(QDialog):
    def __init__(self, settings: QSettings, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.settings = settings
        self.setWindowTitle(tr("title_advanced_settings"))
        self.resize(760, 680)

        root = QVBoxLayout(self)

        xlsx_box = QGroupBox(tr("group_xlsx_export"))
        xlsx_layout = QVBoxLayout(xlsx_box)
        xlsx_layout.addWidget(QLabel(tr("label_xlsx_columns")))
        self.columns_list = QListWidget()
        self.columns_list.setSelectionMode(QAbstractItemView.NoSelection)
        current_columns = normalize_export_headers(self._settings_columns())
        for header in INDEX_HEADERS:
            item = QListWidgetItem(header)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            item.setCheckState(Qt.Checked if header in current_columns else Qt.Unchecked)
            self.columns_list.addItem(item)
        xlsx_layout.addWidget(self.columns_list, 1)

        col_buttons = QHBoxLayout()
        select_all_btn = QPushButton(tr("button_select_all"))
        select_all_btn.clicked.connect(self.select_all_columns)
        default_cols_btn = QPushButton(tr("button_select_default"))
        default_cols_btn.clicked.connect(self.select_default_columns)
        col_buttons.addWidget(select_all_btn)
        col_buttons.addWidget(default_cols_btn)
        col_buttons.addStretch(1)
        xlsx_layout.addLayout(col_buttons)
        root.addWidget(xlsx_box, 2)

        filename_box = QGroupBox(tr("group_filename"))
        filename_form = QFormLayout(filename_box)
        self.filename_template_edit = QLineEdit(str(self.settings.value("filename_template", DEFAULT_FILENAME_TEMPLATE)))
        self.filename_max_spin = QSpinBox()
        self.filename_max_spin.setRange(40, 240)
        self.filename_max_spin.setValue(clamp_filename_max_length(self.settings.value("filename_max_length", DEFAULT_FILENAME_MAX_LENGTH)))
        fields_text = ", ".join("{" + name + "}" for name, _source in FILENAME_TEMPLATE_FIELDS)
        fields_label = QLabel(fields_text)
        fields_label.setWordWrap(True)
        filename_form.addRow(tr("label_filename_template"), self.filename_template_edit)
        filename_form.addRow(tr("label_filename_max_length"), self.filename_max_spin)
        filename_form.addRow(tr("label_available_filename_fields"), fields_label)
        root.addWidget(filename_box)

        log_box = QGroupBox(tr("group_log_file"))
        log_form = QFormLayout(log_box)
        self.save_log_check = QCheckBox(tr("check_save_log_file"))
        self.save_log_check.setChecked(str(self.settings.value("save_log_file", "false")).lower() == "true")
        self.log_dir_edit = QLineEdit(str(self.settings.value("log_dir", self.default_log_dir())))
        self.log_dir_btn = QPushButton(tr("button_choose"))
        self.log_dir_btn.clicked.connect(self.choose_log_dir)
        log_dir_row = QHBoxLayout()
        log_dir_row.addWidget(self.log_dir_edit, 1)
        log_dir_row.addWidget(self.log_dir_btn)
        log_form.addRow("", self.save_log_check)
        log_form.addRow(tr("label_log_folder"), log_dir_row)
        root.addWidget(log_box)

        bottom = QHBoxLayout()
        defaults_btn = QPushButton(tr("button_defaults"))
        defaults_btn.clicked.connect(self.restore_defaults)
        bottom.addWidget(defaults_btn)
        bottom.addStretch(1)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        bottom.addWidget(button_box)
        root.addLayout(bottom)

    def default_log_dir(self) -> str:
        base = str(self.settings.value("backup_dir", str(Path.home() / "PEC_Backup"))).strip()
        try:
            return str(Path(base).expanduser() / "logs")
        except Exception:
            return str(Path.home() / "PEC_Backup" / "logs")

    def _settings_columns(self) -> List[str]:
        raw = str(self.settings.value("xlsx_columns", "")).strip()
        if not raw:
            return list(DEFAULT_XLSX_EXPORT_HEADERS)
        return [x.strip() for x in raw.splitlines() if x.strip()]

    def selected_columns(self) -> List[str]:
        columns: List[str] = []
        for i in range(self.columns_list.count()):
            item = self.columns_list.item(i)
            if item.checkState() == Qt.Checked:
                columns.append(item.text())
        return normalize_export_headers(columns)

    def select_all_columns(self) -> None:
        for i in range(self.columns_list.count()):
            self.columns_list.item(i).setCheckState(Qt.Checked)

    def select_default_columns(self) -> None:
        defaults = set(DEFAULT_XLSX_EXPORT_HEADERS)
        for i in range(self.columns_list.count()):
            item = self.columns_list.item(i)
            item.setCheckState(Qt.Checked if item.text() in defaults else Qt.Unchecked)

    def restore_defaults(self) -> None:
        self.select_default_columns()
        self.filename_template_edit.setText(DEFAULT_FILENAME_TEMPLATE)
        self.filename_max_spin.setValue(DEFAULT_FILENAME_MAX_LENGTH)
        self.save_log_check.setChecked(False)
        self.log_dir_edit.setText(self.default_log_dir())

    def choose_log_dir(self) -> None:
        d = QFileDialog.getExistingDirectory(self, tr("dialog_choose_log_folder"), self.log_dir_edit.text() or self.default_log_dir())
        if d:
            self.log_dir_edit.setText(d)

    def save_to_settings(self) -> None:
        self.settings.setValue("xlsx_columns", "\n".join(self.selected_columns()))
        self.settings.setValue("filename_template", self.filename_template_edit.text().strip() or DEFAULT_FILENAME_TEMPLATE)
        self.settings.setValue("filename_max_length", str(clamp_filename_max_length(self.filename_max_spin.value())))
        self.settings.setValue("save_log_file", "true" if self.save_log_check.isChecked() else "false")
        self.settings.setValue("log_dir", self.log_dir_edit.text().strip() or self.default_log_dir())
        self.settings.sync()


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.settings = QSettings(APP_ORG, APP_NAME)
        set_language(str(self.settings.value("language", "it")))
        if APP_ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_PATH)))
        self.setWindowTitle(tr("app_title"))
        self.resize(1180, 760)
        self.current_thread: Optional[QThread] = None
        self.current_worker: Optional[QObject] = None
        self._worker_done_slot = None
        self._worker_result: Any = None
        self._worker_error = ""
        self._worker_cancelled = ""
        self.pending_delete_candidates: List[Dict[str, str]] = []
        self.pending_index_path = ""
        self.preview_rows: List[Dict[str, Any]] = []
        self._log_file_path: Optional[Path] = None
        self._log_file_error_reported = False
        self._build_ui()
        self._load_settings()

    def _build_ui(self) -> None:
        root = QWidget()
        self.setCentralWidget(root)
        main = QVBoxLayout(root)

        self.conn_box = QGroupBox()
        conn_form = QFormLayout(self.conn_box)
        self.host_edit = QLineEdit()
        self.port_spin = QSpinBox()
        self.port_spin.setRange(1, 65535)
        self.port_spin.setValue(993)
        self.timeout_spin = QSpinBox()
        self.timeout_spin.setRange(5, 600)
        self.timeout_spin.setValue(60)
        self.timeout_spin.setSuffix(" s")
        self.ssl_combo = QComboBox()
        self.user_edit = QLineEdit()
        self.pass_edit = QLineEdit()
        self.pass_edit.setEchoMode(QLineEdit.Password)
        self.language_combo = QComboBox()
        self.save_pass_check = QCheckBox()
        self.debug_check = QCheckBox()

        self.host_label = QLabel()
        self.port_label = QLabel()
        self.timeout_label = QLabel()
        self.security_label = QLabel()
        self.user_label = QLabel()
        self.password_label = QLabel()
        self.language_label = QLabel()
        conn_form.addRow(self.host_label, self.host_edit)
        conn_form.addRow(self.port_label, self.port_spin)
        conn_form.addRow(self.timeout_label, self.timeout_spin)
        conn_form.addRow(self.security_label, self.ssl_combo)
        conn_form.addRow(self.user_label, self.user_edit)
        conn_form.addRow(self.password_label, self.pass_edit)
        conn_form.addRow(self.language_label, self.language_combo)
        conn_form.addRow("", self.save_pass_check)
        conn_form.addRow("", self.debug_check)

        self.opt_box = QGroupBox()
        opt_layout = QGridLayout(self.opt_box)
        self.dir_edit = QLineEdit()
        self.dir_btn = QPushButton()
        self.dir_btn.clicked.connect(self.choose_dir)
        self.cutoff_date = QDateEdit()
        self.cutoff_date.setCalendarPopup(True)
        self.cutoff_date.setDisplayFormat("yyyy-MM-dd")
        self.cutoff_date.setDate(QDate.currentDate())
        self.size_check = QCheckBox()
        self.size_value_spin = QDoubleSpinBox()
        self.size_value_spin.setRange(0.0, 999999.0)
        self.size_value_spin.setDecimals(2)
        self.size_value_spin.setValue(0.0)
        self.size_unit_combo = QComboBox()
        self.size_unit_combo.addItems(["B", "KB", "MB", "GB"])
        size_row = QHBoxLayout()
        size_row.addWidget(self.size_check)
        size_row.addWidget(self.size_value_spin)
        size_row.addWidget(self.size_unit_combo)
        size_row.addStretch(1)

        self.preview_deep_check = QCheckBox()
        self.preview_deep_check.setChecked(True)
        self.delete_check = QCheckBox()
        self.delete_check.setChecked(True)
        self.expunge_check = QCheckBox()
        self.expunge_check.setChecked(False)
        self.folders_list = QListWidget()
        self.folders_list.setFixedHeight(130)
        self.folders_list.setSelectionMode(QAbstractItemView.MultiSelection)
        self.populate_folders(["INBOX"], ["INBOX"])

        self.local_folder_label = QLabel()
        self.cutoff_label = QLabel()
        self.folders_label = QLabel()
        self.folders_help_label = QLabel()
        opt_layout.addWidget(self.local_folder_label, 0, 0)
        opt_layout.addWidget(self.dir_edit, 0, 1)
        opt_layout.addWidget(self.dir_btn, 0, 2)
        opt_layout.addWidget(self.cutoff_label, 1, 0)
        opt_layout.addWidget(self.cutoff_date, 1, 1)
        opt_layout.addLayout(size_row, 2, 1, 1, 2)
        opt_layout.addWidget(self.folders_label, 3, 0, Qt.AlignTop)
        opt_layout.addWidget(self.folders_list, 3, 1, 1, 2)
        opt_layout.addWidget(self.folders_help_label, 4, 1, 1, 2)
        opt_layout.addWidget(self.preview_deep_check, 5, 1, 1, 2)
        opt_layout.addWidget(self.delete_check, 6, 1, 1, 2)
        opt_layout.addWidget(self.expunge_check, 7, 1, 1, 2)

        top = QHBoxLayout()
        top.addWidget(self.conn_box, 1)
        top.addWidget(self.opt_box, 2)
        main.addLayout(top)

        btns = QHBoxLayout()
        self.save_settings_btn = QPushButton()
        self.advanced_settings_btn = QPushButton()
        self.list_btn = QPushButton()
        self.preview_btn = QPushButton()
        self.backup_btn = QPushButton()
        self.stop_btn = QPushButton()
        self.stop_btn.setEnabled(False)
        self.save_settings_btn.clicked.connect(self._save_settings)
        self.advanced_settings_btn.clicked.connect(self.open_advanced_settings)
        self.list_btn.clicked.connect(self.list_folders)
        self.preview_btn.clicked.connect(self.preview)
        self.backup_btn.clicked.connect(self.backup)
        self.stop_btn.clicked.connect(self.stop_current_worker)
        btns.addWidget(self.save_settings_btn)
        btns.addWidget(self.advanced_settings_btn)
        btns.addWidget(self.list_btn)
        btns.addStretch(1)
        btns.addWidget(self.preview_btn)
        btns.addWidget(self.backup_btn)
        btns.addWidget(self.stop_btn)
        main.addLayout(btns)

        self.table = QTableWidget(0, 8)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.cellDoubleClicked.connect(self.open_mail_source)

        lower_box = QWidget()
        lower_layout = QVBoxLayout(lower_box)
        lower_layout.setContentsMargins(0, 0, 0, 0)
        self.progress = QProgressBar()
        self.progress.setValue(0)
        lower_layout.addWidget(self.progress)

        log_header = QHBoxLayout()
        self.log_label = QLabel()
        log_header.addWidget(self.log_label)
        log_header.addStretch(1)
        self.clear_log_btn = QPushButton()
        self.clear_log_btn.clicked.connect(self.clear_logs)
        log_header.addWidget(self.clear_log_btn)
        lower_layout.addLayout(log_header)

        self.log_edit = QPlainTextEdit()
        self.log_edit.setReadOnly(True)
        self.log_edit.setMaximumBlockCount(5000)
        lower_layout.addWidget(self.log_edit)

        self.main_splitter = QSplitter(Qt.Vertical)
        self.main_splitter.addWidget(self.table)
        self.main_splitter.addWidget(lower_box)
        self.main_splitter.setStretchFactor(0, 4)
        self.main_splitter.setStretchFactor(1, 1)
        self.main_splitter.setSizes([520, 180])
        main.addWidget(self.main_splitter, 1)

        self._refresh_language_combo()
        self._refresh_ssl_combo()
        self.language_combo.currentIndexChanged.connect(self.on_language_changed)
        self.retranslate_ui()

    def _refresh_language_combo(self) -> None:
        current = CURRENT_LANGUAGE
        self.language_combo.blockSignals(True)
        self.language_combo.clear()
        for code, label in SUPPORTED_LANGUAGES.items():
            self.language_combo.addItem(label, code)
        idx = self.language_combo.findData(current)
        if idx >= 0:
            self.language_combo.setCurrentIndex(idx)
        self.language_combo.blockSignals(False)

    def _refresh_ssl_combo(self) -> None:
        current = self.ssl_combo.currentData() or self.ssl_combo.currentText() or "SSL/TLS"
        self.ssl_combo.blockSignals(True)
        self.ssl_combo.clear()
        self.ssl_combo.addItem("SSL/TLS", "SSL/TLS")
        self.ssl_combo.addItem("STARTTLS", "STARTTLS")
        self.ssl_combo.addItem(tr("ssl_none"), "Nessuna")
        idx = self.ssl_combo.findData(current)
        if idx < 0:
            idx = self.ssl_combo.findText(str(current))
        self.ssl_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.ssl_combo.blockSignals(False)

    def current_ssl_mode(self) -> str:
        return str(self.ssl_combo.currentData() or self.ssl_combo.currentText())

    def current_language(self) -> str:
        return str(self.language_combo.currentData() or CURRENT_LANGUAGE)

    def retranslate_ui(self) -> None:
        self.setWindowTitle(tr("app_title"))
        self.conn_box.setTitle(tr("group_connection"))
        self.opt_box.setTitle(tr("group_backup"))
        self.host_label.setText(tr("label_host"))
        self.port_label.setText(tr("label_port"))
        self.timeout_label.setText(tr("label_timeout"))
        self.security_label.setText(tr("label_security"))
        self.user_label.setText(tr("label_user"))
        self.password_label.setText(tr("label_password"))
        self.language_label.setText(tr("label_language"))
        self.save_pass_check.setText(tr("check_save_password"))
        self.debug_check.setText(tr("check_debug_imap"))
        self.local_folder_label.setText(tr("label_local_folder"))
        self.dir_btn.setText(tr("button_choose"))
        self.cutoff_label.setText(tr("label_cutoff"))
        self.size_check.setText(tr("check_size_filter"))
        self.folders_label.setText(tr("label_imap_folders"))
        self.folders_help_label.setText(tr("help_load_folders"))
        self.preview_deep_check.setText(tr("check_deep_pec_preview"))
        self.preview_deep_check.setToolTip(tr("tooltip_deep_pec_preview"))
        self.delete_check.setText(tr("check_deleted_at_end"))
        self.expunge_check.setText(tr("check_expunge"))
        self.save_settings_btn.setText(tr("button_save_settings"))
        self.advanced_settings_btn.setText(tr("button_advanced_settings"))
        self.list_btn.setText(tr("button_test_folders"))
        self.preview_btn.setText(tr("button_preview"))
        self.backup_btn.setText(tr("button_backup"))
        self.stop_btn.setText(tr("button_stop"))
        self.log_label.setText(tr("label_log"))
        self.clear_log_btn.setText(tr("button_clear_log"))
        self.table.setHorizontalHeaderLabels([
            tr("table_folder"),
            tr("table_uid"),
            tr("table_date"),
            tr("table_true_sender"),
            tr("table_from_header"),
            tr("table_to"),
            tr("table_subject"),
            tr("table_size"),
        ])
        self._refresh_ssl_combo()

    @Slot()
    def on_language_changed(self) -> None:
        language = self.current_language()
        set_language(language)
        self.settings.setValue("language", language)
        self.settings.sync()
        self.retranslate_ui()
        self.log(tr("language_changed", language=SUPPORTED_LANGUAGES.get(language, language)))

    def choose_dir(self) -> None:
        d = QFileDialog.getExistingDirectory(self, tr("dialog_choose_backup_folder"), self.dir_edit.text() or str(Path.home()))
        if d:
            self.dir_edit.setText(d)

    def populate_folders(self, folders: Sequence[str], selected: Optional[Sequence[str]] = None) -> None:
        selected_set = set(selected or [])
        self.folders_list.clear()
        for folder in folders:
            item = QListWidgetItem(folder)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            item.setCheckState(Qt.Checked if folder in selected_set else Qt.Unchecked)
            self.folders_list.addItem(item)
        if self.folders_list.count() and not self.selected_folders():
            # Conservative default: select INBOX when present, otherwise the first folder.
            for i in range(self.folders_list.count()):
                item = self.folders_list.item(i)
                if item.text().upper() == "INBOX":
                    item.setCheckState(Qt.Checked)
                    return
            self.folders_list.item(0).setCheckState(Qt.Checked)

    def selected_folders(self) -> List[str]:
        folders: List[str] = []
        for i in range(self.folders_list.count()):
            item = self.folders_list.item(i)
            if item.checkState() == Qt.Checked:
                folders.append(item.text())
        return folders

    def all_folders_in_widget(self) -> List[str]:
        return [self.folders_list.item(i).text() for i in range(self.folders_list.count())]

    def size_filter_bytes(self) -> int:
        if not self.size_check.isChecked():
            return 0
        value = float(self.size_value_spin.value())
        multipliers = {"B": 1, "KB": 1024, "MB": 1024 ** 2, "GB": 1024 ** 3}
        unit = self.size_unit_combo.currentText()
        size = int(value * multipliers.get(unit, 1))
        return max(size, 0)

    def _password_key(self) -> bytes:
        seed = (self.user_edit.text() + "@" + self.host_edit.text() + APP_NAME).encode("utf-8", errors="replace")
        return hashlib.sha256(seed).digest()

    def _encode_password(self, pwd: str) -> str:
        data = pwd.encode("utf-8")
        key = self._password_key()
        x = bytes(b ^ key[i % len(key)] for i, b in enumerate(data))
        return base64.b64encode(x).decode("ascii")

    def _decode_password(self, token: str) -> str:
        try:
            data = base64.b64decode(token.encode("ascii"))
            key = self._password_key()
            raw = bytes(b ^ key[i % len(key)] for i, b in enumerate(data))
            return raw.decode("utf-8")
        except Exception:
            return ""

    def xlsx_export_columns(self) -> List[str]:
        raw = str(self.settings.value("xlsx_columns", "")).strip()
        if not raw:
            return list(DEFAULT_XLSX_EXPORT_HEADERS)
        return normalize_export_headers([x.strip() for x in raw.splitlines() if x.strip()])

    def filename_template(self) -> str:
        return str(self.settings.value("filename_template", DEFAULT_FILENAME_TEMPLATE)).strip() or DEFAULT_FILENAME_TEMPLATE

    def filename_max_length(self) -> int:
        return clamp_filename_max_length(self.settings.value("filename_max_length", DEFAULT_FILENAME_MAX_LENGTH))

    def save_log_file_enabled(self) -> bool:
        return str(self.settings.value("save_log_file", "false")).lower() == "true"

    def configured_log_dir(self) -> Path:
        default = Path(self.dir_edit.text().strip() or str(Path.home() / "PEC_Backup")).expanduser() / "logs"
        raw = str(self.settings.value("log_dir", str(default))).strip()
        return Path(raw or str(default)).expanduser()

    def open_advanced_settings(self) -> None:
        dlg = AdvancedConfigDialog(self.settings, self)
        if dlg.exec() == QDialog.Accepted:
            old_log_dir = self.configured_log_dir()
            old_log_enabled = self.save_log_file_enabled()
            dlg.save_to_settings()
            if old_log_dir != self.configured_log_dir() or old_log_enabled != self.save_log_file_enabled():
                self._log_file_path = None
                self._log_file_error_reported = False
            self.log(tr("log_advanced_settings_saved"))

    def _settings_snapshot(self) -> Dict[str, str]:
        """Editable values to save and verify in QSettings.

        Plain-text password and binary Qt states are excluded, while all visible
        GUI fields are included: connection, cutoff date, size, folders, debug,
        preview/deletion flags and local folder.
        """
        return {
            "host": self.host_edit.text().strip(),
            "port": str(self.port_spin.value()),
            "timeout_seconds": str(self.timeout_spin.value()),
            "ssl_mode": self.current_ssl_mode(),
            "username": self.user_edit.text().strip(),
            "language": self.current_language(),
            "backup_dir": self.dir_edit.text().strip(),
            "cutoff_date": self.cutoff_date.date().toString("yyyy-MM-dd"),
            "folders": "\n".join(self.selected_folders()),
            "available_folders": "\n".join(self.all_folders_in_widget()),
            "size_enabled": "true" if self.size_check.isChecked() else "false",
            "size_value": str(self.size_value_spin.value()),
            "size_unit": self.size_unit_combo.currentText(),
            "preview_deep_pec": "true" if self.preview_deep_check.isChecked() else "false",
            "ask_delete": "true" if self.delete_check.isChecked() else "false",
            "expunge": "true" if self.expunge_check.isChecked() else "false",
            "debug": "true" if self.debug_check.isChecked() else "false",
            "remember_password": "true" if self.save_pass_check.isChecked() else "false",
            "xlsx_columns": "\n".join(self.xlsx_export_columns()),
            "filename_template": self.filename_template(),
            "filename_max_length": str(self.filename_max_length()),
            "save_log_file": "true" if self.save_log_file_enabled() else "false",
            "log_dir": str(self.configured_log_dir()),
        }

    def _restore_qt_state(self) -> None:
        # Qt visual states (geometry/splitter/header) are binary QByteArray values.
        # If the Qt/PySide version or column count changes, they can cause native
        # startup crashes in some environments. For this reason, they are NOT restored
        # automatically; enable restore by starting with --restore-ui-state.
        if not RESTORE_UI_STATE:
            return
        try:
            geometry = self.settings.value("window_geometry")
            if geometry:
                ok = self.restoreGeometry(geometry)
                if not ok:
                    self.log(tr("log_layout_geometry_invalid"))
        except Exception as e:
            self.log(tr("log_layout_restore_geometry_ignored", error=e))
        try:
            splitter_state = self.settings.value("splitter_state")
            if splitter_state:
                ok = self.main_splitter.restoreState(splitter_state)
                if not ok:
                    self.log(tr("log_layout_splitter_invalid"))
        except Exception as e:
            self.log(tr("log_layout_restore_splitter_ignored", error=e))
        try:
            header_state = self.settings.value("table_header_state")
            if header_state:
                ok = self.table.horizontalHeader().restoreState(header_state)
                if not ok:
                    self.log(tr("log_layout_columns_invalid"))
        except Exception as e:
            self.log(tr("log_layout_restore_columns_ignored", error=e))

    def _verify_saved_settings(self, expected: Dict[str, str]) -> List[str]:
        mismatches: List[str] = []
        for key, value in expected.items():
            got = self.settings.value(key, "")
            if str(got) != str(value):
                mismatches.append(key)
        return mismatches

    def _load_settings(self) -> None:
        self.host_edit.setText(self.settings.value("host", ""))
        self.port_spin.setValue(int(self.settings.value("port", 993)))
        self.timeout_spin.setValue(int(self.settings.value("timeout_seconds", 60)))
        language = str(self.settings.value("language", CURRENT_LANGUAGE))
        set_language(language)
        self._refresh_language_combo()
        self.retranslate_ui()

        ssl_mode = self.settings.value("ssl_mode", "SSL/TLS")
        idx = self.ssl_combo.findData(str(ssl_mode))
        if idx < 0:
            idx = self.ssl_combo.findText(str(ssl_mode))
        if idx >= 0:
            self.ssl_combo.setCurrentIndex(idx)
        self.user_edit.setText(self.settings.value("username", ""))
        self.dir_edit.setText(self.settings.value("backup_dir", str(Path.home() / "PEC_Backup")))

        cutoff_raw = str(self.settings.value("cutoff_date", ""))
        if cutoff_raw:
            cutoff_qdate = QDate.fromString(cutoff_raw, "yyyy-MM-dd")
            if cutoff_qdate.isValid():
                self.cutoff_date.setDate(cutoff_qdate)

        selected_raw = str(self.settings.value("folders", "INBOX"))
        selected = [x.strip() for x in selected_raw.splitlines() if x.strip()] or ["INBOX"]
        available_raw = str(self.settings.value("available_folders", "INBOX"))
        available = [x.strip() for x in available_raw.splitlines() if x.strip()] or ["INBOX"]
        # Avoid the old invented "Sent Items" default: folders must come from the real LIST response.
        self.populate_folders(available, selected)

        self.size_check.setChecked(str(self.settings.value("size_enabled", "false")).lower() == "true")
        try:
            self.size_value_spin.setValue(float(self.settings.value("size_value", 0.0)))
        except Exception:
            self.size_value_spin.setValue(0.0)
        size_unit = str(self.settings.value("size_unit", "MB"))
        unit_idx = self.size_unit_combo.findText(size_unit)
        if unit_idx >= 0:
            self.size_unit_combo.setCurrentIndex(unit_idx)

        self.preview_deep_check.setChecked(str(self.settings.value("preview_deep_pec", "true")).lower() == "true")
        self.delete_check.setChecked(str(self.settings.value("ask_delete", "true")).lower() == "true")
        self.expunge_check.setChecked(str(self.settings.value("expunge", "false")).lower() == "true")
        self.debug_check.setChecked(str(self.settings.value("debug", "false")).lower() == "true")
        remember = str(self.settings.value("remember_password", "false")).lower() == "true"
        self.save_pass_check.setChecked(remember)
        if remember:
            self.pass_edit.setText(self._decode_password(str(self.settings.value("password", ""))))

        self._restore_qt_state()

    def _save_settings(self, silent: bool = False) -> None:
        values = self._settings_snapshot()
        for key, value in values.items():
            self.settings.setValue(key, value)

        if self.save_pass_check.isChecked():
            self.settings.setValue("password", self._encode_password(self.pass_edit.text()))
        else:
            self.settings.remove("password")

        # Visual interface state. It is saved with the rest, but is not
        # included in text verification because Qt stores it as QByteArray.
        try:
            self.settings.setValue("window_geometry", self.saveGeometry())
            self.settings.setValue("splitter_state", self.main_splitter.saveState())
            self.settings.setValue("table_header_state", self.table.horizontalHeader().saveState())
        except Exception:
            pass

        self.settings.sync()
        mismatches = self._verify_saved_settings(values)
        if mismatches:
            self.log(tr("log_settings_verify_warning") + ", ".join(mismatches))
        elif not silent:
            self.log(tr("log_settings_saved", count=len(values), cutoff=values["cutoff_date"]))

    def _ensure_log_file_path(self) -> Optional[Path]:
        if not self.save_log_file_enabled():
            return None
        log_dir = self.configured_log_dir()
        if self._log_file_path is None or self._log_file_path.parent != log_dir:
            log_dir.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._log_file_path = log_dir / f"imap_backup_{stamp}.log"
            with self._log_file_path.open("a", encoding="utf-8") as f:
                f.write(f"# {APP_NAME} log\n")
                f.write(f"# script: {Path(__file__).name}\n")
                f.write(f"# started_at: {datetime.now().isoformat(timespec='seconds')}\n\n")
            self.log_edit.appendPlainText(f"[{datetime.now().strftime('%H:%M:%S')}] " + tr("log_file_created", path=str(self._log_file_path)))
        return self._log_file_path

    def log(self, text: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {text}"
        self.log_edit.appendPlainText(line)
        if not self.save_log_file_enabled():
            return
        try:
            path = self._ensure_log_file_path()
            if path is not None:
                with path.open("a", encoding="utf-8") as f:
                    f.write(line + "\n")
        except Exception as e:
            if not self._log_file_error_reported:
                self._log_file_error_reported = True
                self.log_edit.appendPlainText(f"[{datetime.now().strftime('%H:%M:%S')}] " + tr("log_file_write_failed", error=e))

    def clear_logs(self) -> None:
        self.log_edit.clear()

    def set_busy(self, busy: bool) -> None:
        for w in [self.list_btn, self.preview_btn, self.backup_btn, self.save_settings_btn, self.advanced_settings_btn]:
            w.setEnabled(not busy)
        self.stop_btn.setEnabled(busy)

    def read_config(self, require_folders: bool = True) -> RuntimeConfig:
        host = self.host_edit.text().strip()
        username = self.user_edit.text().strip()
        password = self.pass_edit.text()
        backup_dir = Path(self.dir_edit.text().strip()).expanduser()
        folders = self.selected_folders()
        if not host:
            raise ValueError(tr("error_missing_host"))
        if not username:
            raise ValueError(tr("error_missing_user"))
        if not password:
            raise ValueError(tr("error_missing_password"))
        if require_folders and not folders:
            raise ValueError(tr("error_no_folder_selected"))
        qd = self.cutoff_date.date()
        cutoff = datetime(qd.year(), qd.month(), qd.day())
        return RuntimeConfig(
            host=host,
            port=self.port_spin.value(),
            ssl_mode=self.current_ssl_mode(),
            username=username,
            password=password,
            folders=folders,
            cutoff=cutoff,
            min_size_bytes=self.size_filter_bytes(),
            backup_dir=backup_dir,
            ask_delete=self.delete_check.isChecked(),
            expunge=self.expunge_check.isChecked(),
            timeout_seconds=self.timeout_spin.value(),
            debug=self.debug_check.isChecked(),
            preview_deep_pec=self.preview_deep_check.isChecked(),
            xlsx_columns=self.xlsx_export_columns(),
            filename_template=self.filename_template(),
            filename_max_length=self.filename_max_length(),
        )

    def start_worker(self, worker: QObject, done_slot) -> None:
        if self.current_thread is not None and self.current_thread.isRunning():
            QMessageBox.warning(self, tr("title_operation_running"), tr("msg_operation_running"))
            return
        self.set_busy(True)
        self.progress.setValue(0)
        self._worker_done_slot = done_slot
        self._worker_result: Any = None
        self._worker_error = ""
        self._worker_cancelled = ""

        thread = QThread()
        self.current_thread = thread
        self.current_worker = worker
        worker.moveToThread(thread)
        worker.log.connect(self.log)  # type: ignore[attr-defined]
        worker.progress.connect(self.on_progress)  # type: ignore[attr-defined]
        worker.done.connect(self._store_worker_result)  # type: ignore[attr-defined]
        worker.failed.connect(self._store_worker_error)  # type: ignore[attr-defined]
        worker.cancelled.connect(self._store_worker_cancelled)  # type: ignore[attr-defined]
        # Safe thread shutdown sequence:
        # - delete the worker with deleteLater while its event loop is still alive;
        # - then ask the thread to quit;
        # - finalize the GUI result only after QThread.finished.
        # Connecting worker.deleteLater to thread.finished can leave deleteLater without
        # an event loop in the worker thread and, on some PySide/Qt builds, can cause
        # native crashes at the end of the operation.
        worker.done.connect(worker.deleteLater)  # type: ignore[attr-defined]
        worker.failed.connect(worker.deleteLater)  # type: ignore[attr-defined]
        worker.cancelled.connect(worker.deleteLater)  # type: ignore[attr-defined]
        worker.done.connect(thread.quit)  # type: ignore[attr-defined]
        worker.failed.connect(thread.quit)  # type: ignore[attr-defined]
        worker.cancelled.connect(thread.quit)  # type: ignore[attr-defined]
        thread.started.connect(worker.run)  # type: ignore[attr-defined]
        thread.finished.connect(self._on_thread_finished)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def _store_worker_result(self, result: Any) -> None:
        self._worker_result = result

    def _store_worker_error(self, text: str) -> None:
        self._worker_error = text

    def _store_worker_cancelled(self, text: str) -> None:
        self._worker_cancelled = text

    def stop_current_worker(self) -> None:
        if self.current_thread is None or not self.current_thread.isRunning() or self.current_worker is None:
            return
        self.stop_btn.setEnabled(False)
        self.log(tr("log_stop_requested_by_user"))
        worker = self.current_worker
        try:
            # Cooperative call: the worker checks the flag between messages;
            # if blocked in an IMAP command, it also tries to close the socket.
            worker.request_cancel()  # type: ignore[attr-defined]
        except Exception as e:
            self.log(tr("log_stop_request_failed", error=e))

    def _on_thread_finished(self) -> None:
        # Defer finalization to the next GUI event-loop cycle:
        # this gives Qt time to deliver any queued signals already emitted
        # by the worker before QThread.finished. It also reduces native crashes
        # at the end of backup/deletion on some Linux + PySide6 combinations.
        QTimer.singleShot(0, self._finalize_finished_worker)

    def _finalize_finished_worker(self) -> None:
        done_slot = getattr(self, "_worker_done_slot", None)
        result = getattr(self, "_worker_result", None)
        error = getattr(self, "_worker_error", "")
        cancelled = getattr(self, "_worker_cancelled", "")
        self.current_thread = None
        self.current_worker = None
        self._worker_done_slot = None
        self.set_busy(False)
        if cancelled:
            self.on_worker_cancelled(cancelled)
        elif error:
            self.on_worker_failed(error)
        elif done_slot is not None:
            done_slot(result)

    def closeEvent(self, event) -> None:  # type: ignore[override]
        if self.current_thread is not None and self.current_thread.isRunning():
            QMessageBox.warning(
                self,
                tr("title_operation_running"),
                tr("msg_close_while_running"),
            )
            event.ignore()
            return
        self._save_settings(silent=True)
        super().closeEvent(event)

    @Slot(int, int)
    def on_progress(self, done: int, total: int) -> None:
        self.progress.setMaximum(max(total, 1))
        self.progress.setValue(done)

    @Slot(str)
    def on_worker_failed(self, text: str) -> None:
        self.log(tr("prefix_error") + text)
        QMessageBox.critical(self, tr("title_error"), text)

    @Slot(str)
    def on_worker_cancelled(self, text: str) -> None:
        self.log(tr("prefix_stop") + text)
        QMessageBox.information(self, tr("title_operation_cancelled"), text)

    def list_folders(self) -> None:
        try:
            cfg = self.read_config(require_folders=False)
        except Exception as e:
            QMessageBox.warning(self, tr("title_configuration"), str(e))
            return
        if cfg.debug:
            self.log(tr("log_debug_enabled"))
        self.log(tr("log_listing_folders"))
        self.start_worker(ListFoldersWorker(cfg), self.on_list_done)

    @Slot(list)
    def on_list_done(self, folders: List[str]) -> None:
        if folders:
            previous = set(self.selected_folders())
            self.populate_folders(folders, previous)
            self._save_settings()
            self.log(tr("log_folders_loaded") + "\n" + "\n".join(folders))
            QMessageBox.information(
                self,
                tr("title_imap_folders"),
                tr("msg_folders_loaded") + "\n\n" + "\n".join(folders[:200]),
            )
        else:
            self.log(tr("log_no_folders"))

    def preview(self) -> None:
        try:
            cfg = self.read_config()
        except Exception as e:
            QMessageBox.warning(self, tr("title_configuration"), str(e))
            return
        self.table.setRowCount(0)
        if cfg.debug:
            self.log(tr("log_debug_enabled"))
        if cfg.min_size_bytes:
            self.log(tr("log_size_filter_active", size=human_size(cfg.min_size_bytes)))
        if cfg.preview_deep_pec:
            self.log(tr("log_deep_preview_active"))
        self.log(tr("log_preview_running"))
        self.start_worker(PreviewWorker(cfg), self.on_preview_done)

    @Slot(list)
    def on_preview_done(self, rows: List[Dict[str, Any]]) -> None:
        self.preview_rows = rows
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            display_from = row.get("display_from") or row.get("pec_party") or row.get("header_from", "")
            subject = row.get("pec_oggetto") or row.get("subject_header", "")
            values = [
                row.get("folder", ""),
                row.get("uid", ""),
                row.get("date_header", "") or row.get("internal_date", ""),
                display_from,
                row.get("header_from", ""),
                row.get("header_to", ""),
                subject,
                f"{human_size(int(row.get('size_bytes') or 0))} ({row.get('size_bytes') or 0})",
            ]
            for c, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if c == 3 and row.get("preview_mode") == "pec_profonda":
                    item.setToolTip(tr("tooltip_true_sender_source"))
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()
        deep_count = sum(1 for row in rows if row.get("preview_mode") == "pec_profonda")
        extra = tr("log_deep_count", count=deep_count) if deep_count else ""
        self.log(tr("log_preview_done", count=len(rows), extra=extra))

    def open_mail_source(self, row: int, column: int) -> None:
        folder_item = self.table.item(row, 0)
        uid_item = self.table.item(row, 1)
        if folder_item is None or uid_item is None:
            return
        folder = folder_item.text()
        uid = uid_item.text()
        try:
            cfg = self.read_config(require_folders=False)
        except Exception as e:
            QMessageBox.warning(self, tr("title_configuration"), str(e))
            return
        self.log(tr("log_open_source", uid=uid, folder=folder))
        self.start_worker(MailSourceWorker(cfg, folder, uid), self.on_mail_source_done)

    @Slot(dict)
    def on_mail_source_done(self, result: Dict[str, Any]) -> None:
        meta = result.get("meta", {}) or {}
        folder = result.get("folder", "")
        uid = result.get("uid", "")
        true_from = meta.get("pec_party") or meta.get("header_from") or ""
        header_from = meta.get("header_from") or ""
        subject = meta.get("pec_oggetto") or meta.get("subject_header") or ""
        info = tr(
            "mail_source_info",
            folder=folder,
            uid=uid,
            size=human_size(int(result.get("raw_bytes") or 0)),
            true_from=true_from,
            header_from=header_from,
            subject=subject,
        )
        if result.get("truncated"):
            info += "\n" + tr("mail_source_truncated_note")

        dlg = QDialog(self)
        dlg.setWindowTitle(tr("title_message_source", uid=uid))
        dlg.resize(1050, 800)
        layout = QVBoxLayout(dlg)
        label = QLabel(info)
        label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(label)

        source_edit = QPlainTextEdit()
        source_edit.setReadOnly(True)
        source_edit.setPlainText(str(result.get("source", "")))
        layout.addWidget(source_edit, 1)

        close_btn = QPushButton(tr("button_close"))
        close_btn.clicked.connect(dlg.accept)
        row_layout = QHBoxLayout()
        row_layout.addStretch(1)
        row_layout.addWidget(close_btn)
        layout.addLayout(row_layout)
        dlg.exec()

    def backup(self) -> None:
        try:
            cfg = self.read_config()
        except Exception as e:
            QMessageBox.warning(self, tr("title_configuration"), str(e))
            return
        cfg.backup_dir.mkdir(parents=True, exist_ok=True)
        self._save_settings()
        if cfg.debug:
            self.log(tr("log_debug_enabled"))
        if cfg.min_size_bytes:
            self.log(tr("log_size_filter_active", size=human_size(cfg.min_size_bytes)))
        self.log(tr("log_backup_started"))
        self.start_worker(BackupWorker(cfg), self.on_backup_done)

    @Slot(dict)
    def on_backup_done(self, result: Dict[str, Any]) -> None:
        self.pending_delete_candidates = result.get("candidates", [])
        self.pending_index_path = result.get("index_path", "")
        msg = tr(
            "backup_result_message",
            total=result.get("total", 0),
            saved=result.get("saved", 0),
            skipped=result.get("skipped", 0),
            failed=result.get("failed", 0),
            index=result.get("index_path", ""),
        )
        self.log(msg.replace("\n", " | "))
        QMessageBox.information(self, tr("title_backup"), msg)

        if result.get("failed", 0):
            self.log(tr("log_no_delete_due_errors"))
            return
        if not self.delete_check.isChecked():
            self.log(tr("log_delete_not_requested"))
            return
        if not self.pending_delete_candidates:
            self.log(tr("log_no_messages_to_delete"))
            return

        expunge_text = "\n\n" + tr("delete_confirm_expunge_warning") if self.expunge_check.isChecked() else ""
        answer = QMessageBox.question(
            self,
            tr("title_confirm_imap_delete"),
            tr("delete_confirm_message", count=len(self.pending_delete_candidates), expunge=expunge_text),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer == QMessageBox.Yes:
            try:
                cfg = self.read_config()
            except Exception as e:
                QMessageBox.warning(self, tr("title_configuration"), str(e))
                return
            self.log(tr("log_delete_confirmed"))
            self.start_worker(DeleteWorker(cfg, self.pending_delete_candidates, self.pending_index_path), self.on_delete_done)
        else:
            self.log(tr("log_delete_cancelled"))

    @Slot(dict)
    def on_delete_done(self, result: Dict[str, Any]) -> None:
        msg = tr("delete_result_message", deleted=result.get("deleted", 0), failed=result.get("failed", 0))
        if result.get("failures"):
            msg += "\n\n" + "\n".join(result["failures"][:20])
        self.log(msg.replace("\n", " | "))
        # Avoid accidental repeated deletion in the same session.
        self.pending_delete_candidates = []
        QMessageBox.information(self, tr("title_imap_deletion"), msg)


def _install_qt_debug_handler() -> None:
    if not QT_DEBUG_PLUGINS_REQUESTED:
        return

    def handler(mode, context, message):  # type: ignore[no-untyped-def]
        line = f"QT[{mode}] {message}"
        try:
            if getattr(context, "file", None):
                line += f" ({context.file}:{context.line})"
        except Exception:
            pass
        try:
            print(line, file=sys.stderr)
        except Exception:
            pass
        try:
            if _CRASH_LOG_FILE is not None:
                _CRASH_LOG_FILE.write(line + "\n")
        except Exception:
            pass

    qInstallMessageHandler(handler)


def _print_debug_help() -> None:
    print(
        """Debug usage:
  python imap_backup.py --safe-mode
      Start without restoring the Qt layout and force software rendering.

  python imap_backup.py --reset-settings
      Clear saved settings and exit. Useful when crashes depend on corrupted
      QSettings/layout data.

  python imap_backup.py --qt-debug-plugins
      Write Qt/plugin diagnostic messages also to pec_imap_backup_crash.log.

  python imap_backup.py --force-xcb
  python imap_backup.py --force-wayland
      Force the Qt graphics backend on Linux.

  PEC_IMAP_BACKUP_CRASH_LOG=/path/log.txt python imap_backup.py
      Change the crash log path.
"""
    )

def main() -> int:
    if "--debug-help" in EARLY_ARGS:
        _print_debug_help()
        return 0

    _install_qt_debug_handler()
    app = QApplication(clean_qt_argv(sys.argv))

    if RESET_SETTINGS_REQUESTED:
        settings = QSettings(APP_ORG, APP_NAME)
        settings.clear()
        settings.sync()
        print(tr("settings_reset_message", org=APP_ORG, app=APP_NAME))
        return 0

    w = MainWindow()
    if SAFE_MODE:
        w.log(tr("log_safe_mode_active"))
    elif not RESTORE_UI_STATE:
        w.log(tr("log_layout_not_restored"))
    w.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
